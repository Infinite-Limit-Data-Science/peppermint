import base64
import csv
import itertools
import pathlib
import sys
from collections import deque
from typing import Any, Iterable, List, Tuple

import openpyxl

CellCoord = Tuple[int, int]

def workbook_from_path(path: pathlib.Path) -> Tuple[List[openpyxl.worksheet.worksheet.Worksheet], bool]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8") as fh:
            reader = list(csv.x(fh))
        max_col = max((len(r) for r in reader), default=0)

        class _CsvSheet:
            title = path.stem
            max_row = len(reader)
            max_column = max_col

            def iter_rows(self, values_only: bool = True):
                for r in reader:
                    yield [c if c != "" else None for c in r] + [None] * (max_col - len(r))

            _images: list = []

        return [_CsvSheet()], True

    wb = openpyxl.load_workbook(path, data_only=True)
    return wb.worksheets, False


def non_empty_matrix(ws) -> List[List[Any]]:
    data = [[None] * ws.max_column for _ in range(ws.max_row)]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            if val not in (None, "", b""):
                data[r_idx][c_idx] = val
    return data

def bounding_box(cells: Iterable[CellCoord]) -> Tuple[int, int, int, int]:
    rows, cols = zip(*cells)
    return min(rows), min(cols), max(rows), max(cols)


def adjacent(r: int, c: int, max_r: int, max_c: int) -> Iterable[CellCoord]:
    if r > 0:          yield (r - 1, c)
    if r < max_r - 1:  yield (r + 1, c)
    if c > 0:          yield (r, c - 1)
    if c < max_c - 1:  yield (r, c + 1)


def find_tables(grid: List[List[Any]]) -> List[dict]:
    if not grid:
        return []

    max_r, max_c = len(grid), len(grid[0])
    visited = [[False] * max_c for _ in range(max_r)]
    tables = []

    for r, c in itertools.product(range(max_r), range(max_c)):
        if grid[r][c] is None or visited[r][c]:
            continue

        q = deque([(r, c)])
        island: list[CellCoord] = []
        while q:
            rr, cc = q.popleft()
            if visited[rr][cc] or grid[rr][cc] is None:
                continue
            visited[rr][cc] = True
            island.append((rr, cc))
            for nr, nc in adjacent(rr, cc, max_r, max_c):
                if not visited[nr][nc] and grid[nr][nc] is not None:
                    q.append((nr, nc))

        if island:
            r0, c0, r1, c1 = bounding_box(island)

            rows = [
                [grid[rr][cc] for cc in range(c0, c1 + 1)]
                for rr in range(r0, r1 + 1)
            ]
            tables.append(
                {
                    "type": "table",
                    "bbox": [r0, c0, r1, c1],
                    "row_cnt": len(rows),
                    "col_cnt": max((len(r) for r in rows), default=0),
                    "rows": rows,
                    "text": "",
                }
            )
    return tables

def extract_images(ws, sheet_idx: int) -> List[dict]:
    out = []
    for img in getattr(ws, "_images", []):
        try:
            anchor = img.anchor
            if hasattr(anchor, "_from"):
                row0 = anchor._from.row
                col0 = anchor._from.col
                row1 = getattr(anchor, "_to", anchor._from).row
                col1 = getattr(anchor, "_to", anchor._from).col
            else:
                row0 = col0 = row1 = col1 = 0
        except Exception:
            row0 = col0 = row1 = col1 = 0

        try:
            raw = img._data()
        except TypeError:
            raw = img._data

        b64 = base64.b64encode(raw).decode()
        out.append(
            {
                "type": "image",
                "sheet": sheet_idx + 1,
                "name": ws.title,
                "bbox": [row0, col0, row1, col1],
                "width": getattr(img, "width", None),
                "height": getattr(img, "height", None),
                "text": b64,
            }
        )
    return out

def iterate_spreadsheet(file_path: str | pathlib.Path):
    path = pathlib.Path(file_path)
    sheets, is_csv = workbook_from_path(path)

    for s_idx, ws in enumerate(sheets):
        grid = non_empty_matrix(ws)
        for tbl in find_tables(grid):
            tbl.update({"sheet": s_idx + 1, "name": ws.title})
            yield tbl

        if not is_csv:
            for img in extract_images(ws, s_idx):
                yield img

def extract_blocks(xls_path: str | pathlib.Path):
    yield from iterate_spreadsheet(xls_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:  python xlsxcut.py <workbook|csv>", file=sys.stderr)
        sys.exit(1)

    wb_path = pathlib.Path(sys.argv[1]).expanduser()
    blocks = list(iterate_spreadsheet(wb_path))
    import json

    print(json.dumps(blocks, ensure_ascii=False, indent=2))
